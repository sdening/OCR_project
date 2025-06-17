import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook
import fitz
from PIL import Image, ImageEnhance
from transformers import TrOCRProcessor, VisionEncoderDecoderModel
from tqdm import tqdm
import re
import time
import io
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import re
import cv2
import numpy as np
from PIL import Image
#import pytesseract
import io
import easyocr
import tkinter.simpledialog as simpledialog

# Function to get month from week number
def get_month_from_week(year, week):
    first_day_of_week = datetime.strptime(f'{year}-W{int(week)}-1', "%Y-W%W-%w")
    return first_day_of_week.strftime("%B")

# Function to check if time is in valid format HH:MM
def is_valid_time(time_str):
    return bool(re.match(r'^([01]\d|2[0-3]):([0-5]\d)$', time_str))

def extract_kalenderwoche(raw_text):
    # Extract all digit sequences from the raw text
    digit_sequences = re.findall(r'\d+', raw_text)
    
    # Join the sequences to form a single number
    if digit_sequences:
        return int(''.join(digit_sequences))
    return 0  # Return 0 if no digits found

def confirm_kalenderwoche(kalenderwoche, year):
    month = get_month_from_week(year, kalenderwoche)
    return messagebox.askyesno(
        "Kalenderwoche bestätigen",
        f'Kalenderwoche "{kalenderwoche}" erkannt - Monat {month}.\nIst das richtig?'
    )

# Function to process the time text from OCR
def process_time_text(text):
    normalized_text = re.sub(r'[ \-%,.]', ':', text)
    normalized_text = re.sub(r'[^\d:]', '', normalized_text)
    if re.match(r'^\d{1,2}:$', normalized_text):
        normalized_text = normalized_text.zfill(3) + "00"
    digits_only = re.sub(r'[^\d]', '', normalized_text)
    if len(digits_only) == 5:
        normalized_text = f"0{digits_only[1:3]}:{digits_only[3:5]}"
    elif len(digits_only) == 4:
        normalized_text = f"{digits_only[:2]}:{digits_only[2:4]}"
    elif len(digits_only) == 3:
        normalized_text = f"0{digits_only[0]}:{digits_only[1:3]}"
    if re.match(r'^\d{1,2}:\d{2}$', normalized_text):
        parts = normalized_text.split(':')
        hour = parts[0].zfill(2)
        minute = parts[1]
        if minute == "60":
            minute = "00"
        return f"{hour}:{minute}"
    parts = re.findall(r'\d+', normalized_text)
    if len(parts) >= 2:
        hour = parts[0].zfill(2)
        minute = parts[1]
        if len(minute) == 1:
            minute = minute + '0'
        if minute == "60":
            minute = "00"
        return f"{hour}:{minute}"
    return text.strip()

# Function to recognize text from a bounding box in PDF
def recognize_number_in_bbox_old(pdf_path, bbox_rect, meaning, processor, model):
    pdf_document = fitz.open(pdf_path)
    page = pdf_document.load_page(0)
    rect = fitz.Rect(*bbox_rect)
    page.set_cropbox(rect)
    pix = page.get_pixmap()
    img = Image.open(io.BytesIO(pix.tobytes())).convert("RGB")
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2)
    pixel_values = processor(images=img, return_tensors="pt").pixel_values
    generated_ids = model.generate(pixel_values, max_new_tokens=20)
    extract_text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0].strip()
    extract_text = re.sub(r'g', '9', extract_text)  # Replace 'g' with '9' here
    pdf_document.close()
    if len(extract_text) > 6 or not re.search(r'\d', extract_text):
        return ""
    if meaning != "kalenderwoche":
        extract_text = process_time_text(extract_text)
        if extract_text == "00:00":
            return ""
    return extract_text

def recognize_number_in_bbox(pdf_path, bbox_rect, meaning, processor, model):
    pdf_document = fitz.open(pdf_path)
    page = pdf_document.load_page(0)
    rect = fitz.Rect(*bbox_rect)

    # Render only the part of the page defined by the rectangle
    pix = page.get_pixmap(clip=rect)
    img = Image.open(io.BytesIO(pix.tobytes())).convert("RGB")
    enhancer = ImageEnhance.Contrast(img)
    img = enhancer.enhance(2)
    pixel_values = processor(images=img, return_tensors="pt").pixel_values
    generated_ids = model.generate(pixel_values, max_new_tokens=20)
    extract_text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0].strip()
    extract_text = re.sub(r'g', '9', extract_text)  # Replace 'g' with '9' here
    pdf_document.close()
    print("Extracted Text: ", extract_text)
    if len(extract_text) > 6 or not re.search(r'\d', extract_text):
        return ""
    if meaning != "kalenderwoche":
        extract_text = process_time_text(extract_text)
        print("\tProcessed Text: ", extract_text)
        if extract_text == "00:00" or re.match(r'^0+$', extract_text.replace(':', '')):
                return ""
    return extract_text


# Function to fill Excel file with recognized data
def fill_excel_from_dict_old(path, data_dict):
    # Load the workbook as a macro-enabled workbook (keep_vba=True to retain macros)
    workbook = load_workbook(path, keep_vba=True)
  
    sheet = workbook["01"]
    year = sheet['D3'].value
    kalenderwoche = int(data_dict.get("kalenderwoche", "").strip('.'))
    month = get_month_from_week(year, kalenderwoche)
    sheet_number = {
        "January": "01",
        "February": "02",
        "March": "03",
        "April": "04",
        "May": "05",
        "June": "06",
        "July": "07",
        "August": "08",
        "September": "09",
        "October": "10",
        "November": "11",
        "December": "12"
    }.get(month, "04")
    sheet_name = f"{sheet_number}"
    sheet = workbook[sheet_name]
    target_row = None
    for row in sheet.iter_rows(min_col=1, max_col=1):
        if row[0].value == kalenderwoche:
            target_row = row[0].row
            break
    if target_row is None:
        messagebox.showerror("Error", f"Kalenderwoche {kalenderwoche} kann nicht in Spalte A gefunden werden.")
        return
    base_row = target_row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    columns = {
        "start": "D",
        "ende": "K",
        "pause_1_start": "E",
        "pause_1_ende": "F",
        "pause_2_start": "G",
        "pause_2_ende": "H",
    }
    days = ["montag", "dienstag", "mittwoch", "donnerstag", "freitag", "samstag", "sonntag"]
    for i, day in enumerate(days):
        row = base_row + i
        for key_suffix, col in columns.items():
            key = f"{key_suffix}_{day}"
            time_value = data_dict.get(key, "")
            cell = sheet[f"{col}{row}"]
            if time_value or time_value == "":
                fill_color = red_fill if time_value and not is_valid_time(time_value) else yellow_fill
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left_cell.value = time_value
                        top_left_cell.fill = fill_color
                        break
                    else:
                        cell.value = time_value
                        cell.fill = fill_color

    # Save the workbook directly as .xlsm, overwriting the original file
    workbook.save(path)

    return path

def fill_excel_from_dict(path, data_dict):
    workbook = load_workbook(path, keep_vba=True)
  
    sheet = workbook["01"]
    year = sheet['D3'].value
    kalenderwoche_raw = data_dict.get("kalenderwoche", "")

    # Extract the Kalenderwoche as an integer
    kalenderwoche = extract_kalenderwoche(kalenderwoche_raw)

    # If the extracted kalenderwoche is not valid, prompt the user
    if not (1 <= kalenderwoche <= 52):
        messagebox.showerror(
            "Error", 
            f"Kalenderwoche ist keine gültige Zahl zwischen 1 und 52. Erkannte Zahl: {kalenderwoche_raw}"
        )
        # Prompt user to enter a valid Kalenderwoche manually
        kalenderwoche = simpledialog.askinteger(
            "Manuelle Eingabe", 
            "Gewünschte Kalenderwoche manuell eingeben:",
            minvalue=1,
            maxvalue=52
        )
        if kalenderwoche is None:
            # If the user cancels the input, stop further processing
            messagebox.showinfo("Abgebrochen", "Der Vorgang wurde abgebrochen.")
            return

    # Confirm the detected kalenderwoche and month with the user
    if not confirm_kalenderwoche(kalenderwoche, year):
        # Prompt user to enter a valid Kalenderwoche manually
        kalenderwoche = simpledialog.askinteger(
            "Manuelle Eingabe", 
            "Gewünschte Kalenderwoche manuell eingeben:",
            minvalue=1,
            maxvalue=52
        )
        if kalenderwoche is None:
            # If the user cancels the input, stop further processing
            messagebox.showinfo("Abgebrochen", "Der Vorgang wurde abgebrochen.")
            return

    month = get_month_from_week(year, kalenderwoche)
    sheet_number = {
        "January": "01",
        "February": "02",
        "March": "03",
        "April": "04",
        "May": "05",
        "June": "06",
        "July": "07",
        "August": "08",
        "September": "09",
        "October": "10",
        "November": "11",
        "December": "12"
    }.get(month, "04")
    sheet_name = f"{sheet_number}"
    sheet = workbook[sheet_name]
    target_row = None
    for row in sheet.iter_rows(min_col=1, max_col=1):
        if row[0].value == kalenderwoche:
            target_row = row[0].row
            break
    if target_row is None:
        messagebox.showerror("Error", f"Kalenderwoche {kalenderwoche} kann nicht in Spalte A gefunden werden.")
        return
    base_row = target_row
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    columns = {
        "start": "D",
        "ende": "K",
        "pause_1_start": "E",
        "pause_1_ende": "F",
        "pause_2_start": "G",
        "pause_2_ende": "H",
    }
    days = ["montag", "dienstag", "mittwoch", "donnerstag", "freitag", "samstag", "sonntag"]
    for i, day in enumerate(days):
        row = base_row + i
        for key_suffix, col in columns.items():
            key = f"{key_suffix}_{day}"
            time_value = data_dict.get(key, "")
            cell = sheet[f"{col}{row}"]
            if time_value or time_value == "":
                fill_color = red_fill if time_value and not is_valid_time(time_value) else yellow_fill
                for merged_range in sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range:
                        top_left_cell = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                        top_left_cell.value = time_value
                        top_left_cell.fill = fill_color
                        break
                    else:
                        cell.value = time_value
                        cell.fill = fill_color

    # Save the workbook directly as .xlsm, overwriting the original file
    workbook.save(path)

    return path

# Function to load the OCR model
def load_model(label, root):
    label.config(text="Model für Handschrifterkennung wird geladen...")
    root.update_idletasks()  # Update the UI
    time.sleep(1)
    
    processor = TrOCRProcessor.from_pretrained('microsoft/trocr-base-handwritten')
    model = VisionEncoderDecoderModel.from_pretrained('microsoft/trocr-base-handwritten')
    
    return processor, model

# Function to process the PDF and Excel files
def process_files_old(processor, model, root, label):
    folder_path = os.getcwd()
    pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]
    if not pdf_files:
        messagebox.showinfo("Keine PDFs gefunden", "Keine PDF-Dateien im Ordner gefunden.")
        return
    pdf_path = os.path.join(folder_path, pdf_files[0])
    data_dict = {}
    bounding_boxes = {
        "kalenderwoche": (450, 700, 700, 770),
        "start_montag": (510, 1160, 730, 1250) ,
        "ende_montag": (510, 1270, 730, 1370),
        "start_dienstag": (510, 1380, 730, 1480),
        "ende_dienstag": (510, 1490, 730, 1580), 
        "start_mittwoch": (510, 1600, 730, 1690),
        "ende_mittwoch": (510, 1710, 730, 1800), 
        "start_donnerstag": (510, 1815, 730, 1915),
        "ende_donnerstag": (510, 1925, 730, 2020), 
        "start_freitag": (510, 2038, 730, 2130),
        "ende_freitag": (510, 2145, 730, 2240),
        "start_samstag": (510, 2260, 730, 2350),
        "ende_samstag":(510, 2365, 730, 2465),

        "pause_1_start_montag": (810, 1160, 980, 1250),
        "pause_1_ende_montag": (810, 1270, 980, 1370),
        "pause_1_start_dienstag": (810, 1380, 980, 1480),
        "pause_1_ende_dienstag": (810, 1490, 980, 1580),
        "pause_1_start_mittwoch": (810, 1600, 980, 1690),
        "pause_1_ende_mittwoch": (810, 1710, 980, 1800),
        "pause_1_start_donnerstag": (810, 1815, 980, 1915),
        "pause_1_ende_donnerstag": (810, 1925, 980, 2020),
        "pause_1_start_freitag": (810, 2038, 980, 2130),
        "pause_1_ende_freitag": (810, 2145, 980, 2240),
        "pause_1_start_samstag": (810, 2260, 980, 2350),
        "pause_1_ende_samstag": (810, 2365, 980, 2465),

        "pause_2_start_montag": (1055, 1160, 1210, 1250),
        "pause_2_ende_montag": (1055, 1270, 1210, 1370),
        "pause_2_start_dienstag": (1055, 1380, 1210, 1480),
        "pause_2_ende_dienstag": (1055, 1490, 1210, 1580),
        "pause_2_start_mittwoch": (1055, 1600, 1210, 1690),
        "pause_2_ende_mittwoch": (1055, 1710, 1210, 1800),
        "pause_2_start_donnerstag": (1055, 1815, 1210, 1915),
        "pause_2_ende_donnerstag": (1055, 1925, 1210, 2020),
        "pause_2_start_freitag": (1055, 2038, 1210, 2130),
        "pause_2_ende_freitag": (1055, 2145, 1210, 2240),
        "pause_2_start_samstag": (1055, 2260, 1210, 2350),
        "pause_2_ende_samstag": (1055, 2365, 1210, 2465)
    }
    for meaning, bbox in bounding_boxes.items():
        recognized_number = recognize_number_in_bbox(pdf_path, bbox, meaning, processor, model)
        data_dict[meaning] = recognized_number
    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsm')]
    
    if not excel_files:
        messagebox.showinfo("Keine Excel-Dateien gefunden", "Keine Excel-Dateien im Ordner gefunden.")
        return
    
    excel_path = os.path.join(folder_path, excel_files[0])
    output_path = fill_excel_from_dict(excel_path, data_dict)
    
    return output_path



    

def process_files(processor, model, root, label):
    folder_path = os.getcwd()
    pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]
    if not pdf_files:
        messagebox.showinfo("Keine PDFs gefunden", "Keine PDF-Dateien im Ordner gefunden.")
        return
    pdf_path = os.path.join(folder_path, pdf_files[0])

    # Preprocess the PDF before processing
    pdf_doc = fitz.open(pdf_path)
    output_doc = fitz.open()
    
    for orig_page in pdf_doc:
        zoom = 3.5
        mat = fitz.Matrix(zoom, zoom)
        pix = orig_page.get_pixmap(matrix=mat)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        img = img.convert('L')
        cleaned_page_array = cv2.adaptiveThreshold(np.array(img), 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 27, 20)
        new_image = Image.fromarray(cleaned_page_array)
        fp = io.BytesIO()  # Pillow will write to this "file pointer"
        new_image.save(fp, "JPEG")  # saving the image in memory
        page = output_doc.new_page(width=new_image.width, height=new_image.height)  # make page in target Document
        page.insert_image(page.rect, stream=fp.getvalue())
    
    processed_pdf_path = pdf_path.replace(".pdf", "_processed.pdf")
    output_doc.save(processed_pdf_path)
    pdf_doc.close()
    output_doc.close()

    # Flags to control the display of status messages
    arbeitszeiten_displayed = False
    pausen_displayed = False

    data_dict = {}
    bounding_boxes = {
        "kalenderwoche": (450, 700, 700, 770),
        "start_montag": (510, 1160, 730, 1250),
        "ende_montag": (510, 1270, 730, 1370),
        "start_dienstag": (510, 1380, 730, 1480),
        "ende_dienstag": (510, 1490, 730, 1580), 
        "start_mittwoch": (510, 1600, 730, 1690),
        "ende_mittwoch": (510, 1710, 730, 1800), 
        "start_donnerstag": (510, 1815, 730, 1915),
        "ende_donnerstag": (510, 1925, 730, 2020), 
        "start_freitag": (510, 2038, 730, 2130),
        "ende_freitag": (510, 2145, 730, 2240),
        "start_samstag": (510, 2260, 730, 2350),
        "ende_samstag": (510, 2365, 730, 2465),

        "pause_1_start_montag": (810, 1160, 980, 1250),
        "pause_1_ende_montag": (810, 1270, 980, 1370),
        "pause_1_start_dienstag": (810, 1380, 980, 1480),
        "pause_1_ende_dienstag": (810, 1490, 980, 1580),
        "pause_1_start_mittwoch": (810, 1600, 980, 1690),
        "pause_1_ende_mittwoch": (810, 1710, 980, 1800),
        "pause_1_start_donnerstag": (810, 1815, 980, 1915),
        "pause_1_ende_donnerstag": (810, 1925, 980, 2020),
        "pause_1_start_freitag": (810, 2038, 980, 2130),
        "pause_1_ende_freitag": (810, 2145, 980, 2240),
        "pause_1_start_samstag": (810, 2260, 980, 2350),
        "pause_1_ende_samstag": (810, 2365, 980, 2465),

        "pause_2_start_montag": (1055, 1160, 1210, 1250),
        "pause_2_ende_montag": (1055, 1270, 1210, 1370),
        "pause_2_start_dienstag": (1055, 1380, 1210, 1480),
        "pause_2_ende_dienstag": (1055, 1490, 1210, 1580),
        "pause_2_start_mittwoch": (1055, 1600, 1210, 1690),
        "pause_2_ende_mittwoch": (1055, 1710, 1210, 1800),
        "pause_2_start_donnerstag": (1055, 1815, 1210, 1915),
        "pause_2_ende_donnerstag": (1055, 1925, 1210, 2020),
        "pause_2_start_freitag": (1055, 2038, 1210, 2130),
        "pause_2_ende_freitag": (1055, 2145, 1210, 2240),
        "pause_2_start_samstag": (1055, 2260, 1210, 2350),
        "pause_2_ende_samstag": (1055, 2365, 1210, 2465)
    }

    # Draw bounding boxes on the processed PDF
    doc = fitz.open(processed_pdf_path)
    page = doc[0]  # Work with the first page for testing
    for meaning, bbox in bounding_boxes.items():
        # Update status messages based on the first occurrence
        if "pause" in meaning and not pausen_displayed:
            label.config(text="Arbeitszeiten werden übertragen... \nPausenzeiten werden übertragen...")
            root.update_idletasks()
            pausen_displayed = True
        elif "pause" not in meaning and not arbeitszeiten_displayed:
            label.config(text="Arbeitszeiten werden übertragen...")
            root.update_idletasks()
            arbeitszeiten_displayed = True

        rect = fitz.Rect(bbox)
        page.draw_rect(rect, color=(1, 0, 0), width=1.5)
    
        recognized_number = recognize_number_in_bbox(processed_pdf_path, bbox, meaning, processor, model)
        data_dict[meaning] = recognized_number
    
    # Save the processed PDF with bounding boxes
    output_pdf_path = processed_pdf_path.replace('_processed.pdf', '_withBB.pdf')
    doc.save(output_pdf_path)
    doc.close()

    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsm')]
    
    if not excel_files:
        messagebox.showinfo("Keine Excel-Dateien gefunden", "Keine Excel-Dateien im Ordner gefunden.")
        return
    
    excel_path = os.path.join(folder_path, excel_files[0])
    output_path = fill_excel_from_dict(excel_path, data_dict)
    
    return output_path


# Function to start the process and handle the UI
def start_process_old(label, root, start_button):
    try:
        # Deactivate the start button to prevent re-clicks
        start_button.config(state=tk.DISABLED)
        
        # Update the UI to indicate that the model is being loaded
        processor, model = load_model(label, root)

        # Once the model is loaded, proceed with processing
        label.config(text="PDF- und Excel-Datei werden verarbeitet...")
        root.update_idletasks()  # Force UI update
        
        output_path = process_files(processor, model, root, label)
        
        label.config(text=f"Fertig!\nNeue Excel ist gespeichert.\n\nFenster schließt in wenigen Sekunden.")
        root.update()  # Ensure all pending updates are processed

        # Close the application immediately after saving the Excel file
        root.after(3000, root.destroy)  # Close the window after 1 second delay to allow updates
    except Exception as e:
        label.config(text=f"Fehler: {str(e)}")
        root.update()  # Ensure error message is shown
        #root.after(1000, root.destroy)  # Close the window after 1 second delay to allow updates

def start_process(label, root, start_button):
    try:
        # Deactivate the start button to prevent re-clicks
        start_button.config(state=tk.DISABLED)
        
        # Select the folder and identify the files
        folder_path = os.getcwd()
        pdf_files = [f for f in os.listdir(folder_path) if f.endswith('.pdf')]
        excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsm')]
        
        if not pdf_files:
            messagebox.showinfo("Keine PDFs gefunden", "Keine PDF-Dateien im Ordner gefunden.")
            return
        if not excel_files:
            messagebox.showinfo("Keine Excel-Dateien gefunden", "Keine Excel-Dateien im Ordner gefunden.")
            return
        
        # Extract file names
        pdf_filename = os.path.basename(pdf_files[0])
        excel_filename = os.path.basename(excel_files[0])

        # Update the UI to show the filenames being processed
        #label.config(text=f"Model für Handschrifterkennung wird geladen.")
        #root.update_idletasks()  # Force UI update

        # Load the OCR model
        processor, model = load_model(label, root)

        # Update the UI to indicate that processing has started
        label.config(text=f"PDF- und Excel-Datei werden verarbeitet...\nPDF: {pdf_filename}\nExcel: {excel_filename}")
        root.update_idletasks()  # Force UI update
        
        # Process the files and get the output path
        output_path = process_files(processor, model, root, label)
        
        # Inform the user and update the UI
        label.config(text=f"Fertig! \nNeue Excel ist gespeichert unter {excel_filename}.\n\nFenster schließt in wenigen Sekunden.")
        root.update()  # Ensure all pending updates are processed

        # Close the application after a short delay
        root.after(3000, root.destroy)  # Close the window after 3 seconds delay to allow updates
    except Exception as e:
        # Handle errors and inform the user
        label.config(text=f"Fehler: {str(e)}")
        root.update()  # Ensure error message is shown




# Main function to run the Tkinter application
def run():
    root = tk.Tk()
    root.title("Arbeitszeitübertragung")
    root.geometry("400x200")

    label = tk.Label(root, text="Klicke ‘Start‘ um die PDF-Datei zu verarbeiten.", font=("Arial", 12))
    label.pack(pady=10)

    start_button = tk.Button(root, text="Start", command=lambda: start_process(label, root, start_button))
    start_button.pack(pady=20)

    root.mainloop()

if __name__ == "__main__":
    run()
